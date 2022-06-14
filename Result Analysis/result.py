import pandas as pd
import tkinter as tk
import openpyxl as xl
import os

from tkinter import filedialog
from tkinter import Menu
from tkinter import messagebox
import re

class Analyser:
    def __init__(self,parent):
        
        self.root = parent
        self.root.title("Main frame")
        self.frame = tk.Frame(parent)
        self.frame.grid()
        menubar = Menu(self.frame)
        menubar1 = Menu(menubar,tearoff = 0)
        menubar1.add_command(label = "New",command = self.new)
        menubar1.add_separator()
        menubar1.add_command(label = "Quit",command = self.root.quit)
        menubar1.add_command(label = "Restart",command = self.restart)
        menubar.add_cascade(label = "File", menu = menubar1)
        tk.Label(self.frame, text = "Make sure the file is csv").grid(row = 1 , column = 0)
        tk.Label(self.frame, text = 'Before using the app make sure the csv is in the format of Result Format.xlsx').grid(row = 2 , column = 0)
        self.root.config(menu = menubar)
    
    def new(self):
        try:
            self.file = filedialog.askopenfile(parent = self.frame,title = "Choose a file",mode = "rb")
            if self.file != None:
                print(self.file.name)
                self.filename = self.src = os.path.split(self.file.name)
                self.filename = self.filename[1]
                self.filename = self.filename
                self.dst1     = self.src[0]
                self.src      = self.src[0]+"/"+self.src[1]
                print(self.src)
                tk.Label(self.frame,text="File is  {filename}".format(filename=self.filename),font=('Helvetica','15')).grid(row=0,column=1)
                if not os.path.isfile(self.file.name):   
                    """dst = file.name + "/{filename}".format(filename=self.filename)
                    cp(self.src,dst)"""
                    self.frame_grid()
                else:
                    self.frame_grid()
        except:
            self.restart(self)
    def frame_grid(self):
        branch = tk.Button(self.frame,text = "Branch Analysis")
        tk.Label(self.frame,text = 'Enter Branch Name :').grid(row = 1 , column = 0)
        bname=tk.Entry(self.frame)
        bname.grid(row = 1 , column = 1)
        bname.focus()
        branch.grid(row = 2 , column = 0)
        ClassName = tk.Button(self.frame, text = "Class Analysis")
        ClassName.bind("<Button-1>",lambda event , bname=bname : self.ClassWise(event,bname))
        ClassName.grid(row = 2, column = 1)
        bname.bind("<Return>",lambda event , bname=bname : self.BranchWise(event,bname))
        branch.bind("<Button-1>", lambda event , bname=bname : self.BranchWise(event,bname)) 
        TopperName = tk.Button(self.frame, text = "Topper's List")
        TopperName.grid(row = 2, column = 2)
        TopperName.bind("<Button-1>", lambda event , bname=bname : self.TopperList(event,bname))
        
    def TopperList(self,event,bname):
        filename    = self.filename
        bname       = bname.get()
        bname=bname.upper()
        sep         = ','
        pd.options.mode.chained_assignment = None
        df = pd.read_csv(self.file.name,sep=sep)
        if bname:
            OverAllTopper = df.sort_values("CGPA",ascending = False)
            OverAllTopper = OverAllTopper.head(10)
            df  = df.query("Branch == @bname")
            df  = df.fillna(-2)
            df  = df.set_index('Sr_No')
            s=df.shape   
            s=s[0]   #Appeared
            if not s == -1:
                BranchTopper = df.sort_values("CGPA",ascending = False)
                BranchTopper = BranchTopper.head(10)
            writer = pd.ExcelWriter("Toppers.xlsx",engine = 'xlsxwriter')
            BranchTopper.to_excel(writer,sheet_name = 'Branch Topper')
            OverAllTopper.to_excel(writer,sheet_name = 'OverAll Topper')
            messagebox.showinfo("File Saved", "File Saved")
            a = re.findall(r"[\w']+",filename)
            self.TableName = ''
            for i in range(len(a)):
                a[i] = a[i] + '_'
                self.TableName += a[i]
            self.TableName += 'TopperList' 
    def ClassWise(self,event, bname):
        try:
            destFile = 'ClassWise Analysis.xlsx'
            #cp('ClassWise Analysis.xlsx', destFile)
            workbook    = xl.load_workbook(destFile)
            sheet = workbook.active
            filename    = self.filename
            bname       = bname.get()
            bname=bname.upper()
            sep         = ','
            pd.options.mode.chained_assignment = None
            df = pd.read_csv(self.file.name,sep=sep)
            a = df.columns.values
            P_F_Columns = []
            INT_Columns = []
            Sub_Columns = []
            T_Columns   = []
            
            a = a[5:]
            for i in range(5):
                Sub_Columns.append(a[5*i])
                P_F_Columns.append(a[5*i+1])
                INT_Columns.append(a[5*i+2])
                T_Columns.append(a[5*i+4])
            if bname:
                df  = df.query("Branch == @bname")
                df  = df.fillna(-2)
                df  = df.set_index('Sr_No')
                s=df.shape   
                s=s[0]   #Appeared
                if not s == -1:
                    Absent = []
                    for col in P_F_Columns:
                        df[col].replace('P', 1, inplace = True)
                        df[col].replace('F', -1, inplace = True)        
                    for col in T_Columns:
                        df[col].replace('#VALUE!', -5, inplace = True)
                    for col in Sub_Columns :
                        df[col].replace("AB", -3, inplace = True)
                    for col in Sub_Columns :
                        int_df      = df[col].astype(int)
                        Absent.append(sum(map(lambda x: x == -3,int_df)))  
                    PassedSubject = []
                    for col in P_F_Columns :
                        int_df      = df[col].astype(int)
                        PassedSubject.append(sum(map(lambda x: x == 1,int_df))) 
                    for k in range(len(Sub_Columns)):
                        sheet.cell(row = 2+k, column  = 1).value = (Sub_Columns[k])
                        sheet.cell(row = 2+k, column  = 2).value = int(s - Absent[k])
                        sheet.cell(row = 2+k, column  = 3).value = int(PassedSubject[k] - Absent[k])
                        sheet.cell(row = 2+k, column  = 4).value = int(s - PassedSubject[k])
                        sheet.cell(row = 2+k, column  = 5).value = int(Absent[k])
                    workbook.save(destFile)
                    messagebox.showinfo("Saved", "File Succesfully Saved")
                    a = re.findall(r"[\w']+",filename)
                    self.TableName = ''
                    for i in range(len(a)):
                        a[i] = a[i] + '_'
                        self.TableName += a[i]
                    self.TableName += 'ClassWiseList' 
        except:
            messagebox.showerror(title = 'Error', message = 'File doesnt exists {}'.format(destFile))
    def BranchWise(self,event,bname):
        try:
            destFile = 'Result analysis.xlsx'
            #cp('Result analysis.xlsx', destFile)
            wb    = xl.load_workbook(destFile)
            sheet = wb.active
            filename    = self.filename
            bname       = bname.get()
            bname=bname.upper()
            sep         = ','
            pd.options.mode.chained_assignment = None
            df = pd.read_csv(self.file.name,sep=sep)
            a = df.columns.values
            P_F_Columns = []
            INT_Columns = []
            Sub_Columns = []
            T_Columns   = []
            
            a = a[5:]
            for i in range(5):
                Sub_Columns.append(a[5*i])
                P_F_Columns.append(a[5*i+1])
                INT_Columns.append(a[5*i+2])
                T_Columns.append(a[5*i+4])
            if bname:
                df  = df.query("Branch == @bname")
                df  = df.fillna(-2)
                df  = df.set_index('Sr_No')
                s=df.shape   
                s=s[0]   #Appeared
                if not s == -1:
                    dict1={}
                    dict2={}
                    dict3={}
                    for col in P_F_Columns:
                        df[col].replace('P', 1, inplace = True)
                        df[col].replace('F', -1, inplace = True)        
                    for col in T_Columns:
                        df[col].replace('#VALUE!', -5, inplace = True)
                    for col in T_Columns:
                        int_df      = df[col].astype(int)
                        below_lvl   = map(lambda x: x < 50, int_df)
                        below_lvl   = (sum(below_lvl)/s)*100
                        between_lvl = map(lambda x: 50 < x < 60, int_df)
                        between_lvl = (sum(between_lvl)/s)*100
                        above_lvl   = map(lambda x: x > 60, int_df)
                        above_lvl   = (sum(above_lvl)/s)*100
                        dict1[col]  = below_lvl 
                        dict2[col]  = between_lvl
                        dict3[col]  = above_lvl  
                    for col in Sub_Columns :
                        df[col].replace("AB", -3, inplace = True)
                    df["External_KT"] = 0
                    df["Internal_KT"] = 0
                    for row in range(df.shape[0]):
                        counter       = 0
                        counter1      = 0 
                        for col in range (len(P_F_Columns)):
                            TheoryMarks   = int(df.iloc[row][Sub_Columns[col]])
                            TotalMarks    = int(df.iloc[row][T_Columns[col]])
                            InternalMarks = int(df.iloc[row][INT_Columns[col]])
                            if df.iloc[row][P_F_Columns[col]] == -1 :
                                if((TheoryMarks >= 32) and (TotalMarks < 40) ):
                                    counter += 1
                                    df.Internal_KT.iloc[row] = counter
                                if((TheoryMarks < 32) and (TotalMarks < 40)  and (InternalMarks < 8)):
                                    counter += 1
                                    df.Internal_KT.iloc[row] = counter 
                                if (TheoryMarks < 32) :
                                    counter1 += 1
                                    df.External_KT.iloc[row] = counter1
                    int_df      = df["GT"].astype(int)
                    failed      = sum(map(lambda x : x == -2, int_df))
                    passed      = s-failed
                    #passpercent = (passed/s)*100
                    failed      = int(failed)
                    passed      = int(passed)
                    ExtKT       = []
                    IntKT       = []
                    for i in range (6):
                        int_df  = df["External_KT"].astype(int)
                        int_df1 = df["Internal_KT"].astype(int) 
                        ExtKT.append(sum(map(lambda x : x == i+1,int_df)))
                        IntKT.append(sum(map(lambda x : x == i+1,int_df1)))
                    j = list(dict1.values())
                    k = list(dict1.keys())
                    for i in range(len(k)):
                        SubjectName = k[i].split("_")
                        SubjectName = SubjectName[1]
                        sheet.cell(row = 6+i, column = 1).value = SubjectName
                        sheet.cell(row = 6+i, column = 3).value = j[i]
                    j = list(dict2.values())
                    k = list(dict1.keys())
                    for i in range(len(k)):
                        sheet.cell(row = 6+i, column = 4).value = j[i]
                    j = list(dict3.values())
                    k = list(dict1.keys())
                    for i in range(len(k)):
                        sheet.cell(row = 6+i, column = 5).value = j[i]
                    sheet.cell(row = 15, column = 1).value = s
                    sheet.cell(row = 15, column = 3).value = failed
                    sheet.cell(row = 15, column = 4).value = passed        
                    for i in range (6):
                        sheet.cell(row = 16, column = 12-i).value = int(ExtKT[i])
                        sheet.cell(row = 15, column = 12-i).value = int(IntKT[i])                                       
                    wb.save(destFile)
                    messagebox.showinfo("Saved", "File Succesfully Saved")
                    a = re.findall(r"[\w']+",filename)
                    self.TableName =''
                    for i in range(len(a)):
                        a[i] = a[i] + '_'
                        self.TableName += a[i]
                elif s == -1:
                    messagebox.showerror("Invalid Branch", "No such Branch exists")
            elif not bname:
                messagebox.showerror("Empty", "No Value Provided")
        except:
            messagebox.showerror(title = 'Error', message = 'File doesnt exists {}'.format(destFile))
    def restart(self):
        self.root.destroy()
        Analyser(root)
if __name__ == "__main__":
    root = tk.Tk()
    app = Analyser(root)
    root.mainloop()